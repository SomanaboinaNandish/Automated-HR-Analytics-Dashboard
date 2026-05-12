"""
HR Automation Dashboard Builder
Senior Business Analyst + Excel Automation Expert Solution
Generates a production-ready HR Dashboard with KPIs, alerts, charts, and analytics
"""

import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill

)
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart.label import DataLabelList
import shutil
import os

# ─────────────────────────────────────────────────────────
#  COLOUR PALETTE  (Dark Corporate Theme)
# ─────────────────────────────────────────────────────────
C = {
    "navy":       "0D1B2A",   # Primary dark background
    "dark_blue":  "1B3A5C",   # Secondary dark
    "mid_blue":   "1E5F8C",   # Accent blue
    "accent":     "2196F3",   # Bright blue accent
    "teal":       "00BCD4",   # Teal highlight
    "white":      "FFFFFF",
    "off_white":  "F4F6F9",
    "light_gray": "E8EDF2",
    "mid_gray":   "B0BEC5",
    "dark_gray":  "546E7A",
    "green":      "00C853",
    "amber":      "FFB300",
    "red":        "EF5350",
    "orange":     "FF7043",
    "purple":     "7E57C2",
    "header_bg":  "0D47A1",
    "row_even":   "F0F4F8",
    "row_odd":    "FAFCFF",
    "kpi_bg1":    "1565C0",
    "kpi_bg2":    "00695C",
    "kpi_bg3":    "4527A0",
    "kpi_bg4":    "B71C1C",
    "alert_red":  "FFCDD2",
    "alert_amber":"FFF9C4",
    "probation":  "E3F2FD",
}

def make_font(name="Arial", size=10, bold=False, italic=False, color="000000"):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border(style="thin"):
    s = Side(style=style, color="B0BEC5")
    return Border(left=s, right=s, top=s, bottom=s)

def make_header_border():
    s = Side(style="medium", color="0D47A1")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def right_align():
    return Alignment(horizontal="right", vertical="center")

def style_header_cell(cell, text, bg=None, fg="FFFFFF", size=10, bold=True):
    cell.value = text
    cell.font = make_font(size=size, bold=bold, color=fg)
    cell.fill = make_fill(bg or C["header_bg"])
    cell.alignment = center()
    cell.border = make_header_border()

def style_data_cell(cell, value, bg="FFFFFF", fg="1A1A1A", bold=False,
                    align="left", size=9, number_format=None):
    cell.value = value
    cell.font = make_font(size=size, bold=bold, color=fg)
    cell.fill = make_fill(bg)
    cell.alignment = center() if align == "center" else (right_align() if align == "right" else left_align())
    cell.border = make_border()
    if number_format:
        cell.number_format = number_format

def freeze_and_filter(ws, freeze_ref, filter_range=None):
    ws.freeze_panes = freeze_ref
    if filter_range:
        ws.auto_filter.ref = filter_range

def set_tab_color(ws, hex_color):
    ws.sheet_properties.tabColor = hex_color

# ─────────────────────────────────────────────────────────
#  LOAD SOURCE DATA
# ─────────────────────────────────────────────────────────
SRC = r"C:\Users\nandi\Desktop\HR\HR_Dashboard_Data.xlsx"

def load_all_sheets():
    xl = pd.read_excel(SRC, sheet_name=None, header=None)
    return xl

def load_india():
    df = pd.read_excel(SRC, sheet_name="India Employee Database")
    df.columns = [
        "EmpID","Name","Department","Designation","ReportingManager",
        "Skillset","DOJ","EmploymentStatus","LWD"
    ]
    df["DOJ"] = pd.to_datetime(df["DOJ"], errors="coerce")
    df["LWD"] = pd.to_datetime(df["LWD"], errors="coerce")
    df["Region"] = "India"
    return df

def load_us():
    df = pd.read_excel(SRC, sheet_name="US Employee Database")
    df.columns = [
        "EmpID","Name","Department","Designation","ReportingManager",
        "Allocation","CTC","Skillset","DOJ","EmploymentStatus","LWD"
    ]
    df["DOJ"] = pd.to_datetime(df["DOJ"], errors="coerce")
    df["LWD"] = pd.to_datetime(df["LWD"], errors="coerce")
    df["Region"] = "US"
    return df

def load_offboarded():
    df = pd.read_excel(SRC, sheet_name="Offboarded Resources")
    df.columns = [
        "EmpID","Name","Department","Region","Designation",
        "DOJ","LWD","ExitReason","ExitQuarter","NoticePeriodServed","RehireEligible"
    ]
    df["DOJ"] = pd.to_datetime(df["DOJ"], errors="coerce")
    df["LWD"] = pd.to_datetime(df["LWD"], errors="coerce")
    return df

def load_risk():
    df = pd.read_excel(SRC, sheet_name="Risk Report")
    df.columns = [
        "EmpID","Name","Department","Region","RiskCategory",
        "RiskLevel","IdentifiedDate","MitigationAction","Status","HRNotes"
    ]
    df["IdentifiedDate"] = pd.to_datetime(df["IdentifiedDate"], errors="coerce")
    return df

def load_productivity():
    df = pd.read_excel(SRC, sheet_name="Productivity")
    cols = ["EmpID","Name","Department","Region"] + \
           [f"M{i}" for i in range(1,13)] + ["OverallAvg","BelowFlag"]
    df.columns = cols
    return df

def load_finance():
    df = pd.read_excel(SRC, sheet_name="Finance")
    df.columns = ["EmpID","Name","Department","Region","AnnualINR","MonthlyINR","AnnualUSD","MonthlyUSD"]
    return df

def load_rm():
    raw = pd.read_excel(SRC, sheet_name="RM Data", header=None)
    # Use row index 1 for proper headers (row 0 is merged header, row 1 has actual sub-cols)
    df = pd.read_excel(SRC, sheet_name="RM Data", header=1)
    return df

# ─────────────────────────────────────────────────────────
#  KPI CALCULATIONS
# ─────────────────────────────────────────────────────────
def compute_kpis(india, us, offboarded, risk, productivity):
    today = pd.Timestamp.today()
    combined = pd.concat([
        india[["EmpID","Name","Department","Designation","EmploymentStatus","LWD","Region","DOJ"]],
        us[["EmpID","Name","Department","Designation","EmploymentStatus","LWD","Region","DOJ"]]
    ], ignore_index=True)

    total_hc       = len(combined)
    india_hc       = len(india)
    us_hc          = len(us)
    confirmed      = (combined["EmploymentStatus"] == "Confirmed").sum()
    on_probation   = (combined["EmploymentStatus"] == "Under Probation").sum()
    interns        = (combined["EmploymentStatus"] == "Intern").sum()

    # Intern LWD alerts (within 45 days)
    intern_alerts = combined[
        (combined["EmploymentStatus"] == "Intern") &
        (combined["LWD"].notna()) &
        (combined["LWD"] - today <= pd.Timedelta(days=45)) &
        (combined["LWD"] >= today)
    ].copy()
    intern_alerts["DaysLeft"] = (intern_alerts["LWD"] - today).dt.days

    # Probation alerts (within 30 days — approximated from DOJ + 90 days)
    combined["ProbationDue"] = combined["DOJ"] + pd.Timedelta(days=90)
    prob_alerts = combined[
        (combined["EmploymentStatus"] == "Under Probation") &
        (combined["ProbationDue"].notna()) &
        (combined["ProbationDue"] - today <= pd.Timedelta(days=30)) &
        (combined["ProbationDue"] >= today)
    ].copy()
    prob_alerts["DaysLeft"] = (prob_alerts["ProbationDue"] - today).dt.days

    # Attrition by quarter
    attrition_q = offboarded.groupby("ExitQuarter").size().reset_index(name="Exits")
    # Approx headcount per quarter (use current as denominator proxy)
    attrition_q["AttritionRate"] = (attrition_q["Exits"] / total_hc * 100).round(2)

    # Department headcount
    dept_hc = combined.groupby("Department").size().reset_index(name="Headcount")
    dept_hc = dept_hc.sort_values("Headcount", ascending=False)

    # Risk distribution
    risk_dist = risk.groupby("RiskLevel").size().reset_index(name="Count")

    # Productivity flags
    below_8 = productivity[productivity["BelowFlag"].notna() & (productivity["BelowFlag"] != "")].shape[0]
    avg_prod = round(productivity["OverallAvg"].mean(), 2)

    # Finance totals
    finance = load_finance()
    total_ctc_inr = finance["AnnualINR"].sum()
    total_ctc_usd = finance["AnnualUSD"].sum()

    return {
        "total_hc": total_hc,
        "india_hc": india_hc,
        "us_hc": us_hc,
        "confirmed": confirmed,
        "on_probation": on_probation,
        "interns": interns,
        "intern_alerts": intern_alerts,
        "prob_alerts": prob_alerts,
        "attrition_q": attrition_q,
        "dept_hc": dept_hc,
        "risk_dist": risk_dist,
        "below_8": below_8,
        "avg_prod": avg_prod,
        "total_ctc_inr": total_ctc_inr,
        "total_ctc_usd": total_ctc_usd,
        "combined": combined,
    }

# ─────────────────────────────────────────────────────────
#  SHEET BUILDERS
# ─────────────────────────────────────────────────────────

def build_india_sheet(wb, india_df):
    if "India Employee Database" in wb.sheetnames:
        del wb["India Employee Database"]
    ws = wb.create_sheet("India Employee Database")
    set_tab_color(ws, C["mid_blue"])

    headers = ["Employee ID","Name","Department","Designation",
               "Reporting Manager","Skillset","Date of Joining",
               "Employment Status","Last Working Day (Interns)"]
    col_widths = [14, 22, 16, 22, 22, 35, 18, 18, 20]

    # Title row
    ws.merge_cells("A1:I1")
    tc = ws["A1"]
    tc.value = "🇮🇳  INDIA EMPLOYEE DATABASE"
    tc.font = make_font(size=14, bold=True, color=C["white"])
    tc.fill = make_fill(C["navy"])
    tc.alignment = center()
    ws.row_dimensions[1].height = 30

    # Header row
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=ci)
        style_header_cell(cell, h, bg=C["header_bg"])
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    today = pd.Timestamp.today()
    for ri, (_, row) in enumerate(india_df.iterrows(), 3):
        bg = C["row_even"] if ri % 2 == 0 else C["row_odd"]
        status = str(row.get("EmploymentStatus",""))
        if status == "Intern":
            lwd = row.get("LWD")
            if pd.notna(lwd) and (lwd - today).days <= 45 and (lwd - today).days >= 0:
                bg = C["alert_red"]
        elif status == "Under Probation":
            bg = C["probation"]

        vals = [
            row["EmpID"], row["Name"], row["Department"], row["Designation"],
            row["ReportingManager"], row["Skillset"],
            row["DOJ"].strftime("%d-%b-%Y") if pd.notna(row["DOJ"]) else "",
            row["EmploymentStatus"],
            row["LWD"].strftime("%d-%b-%Y") if pd.notna(row.get("LWD")) else "N/A"
        ]
        for ci, val in enumerate(vals, 1):
            style_data_cell(ws.cell(row=ri, column=ci), val, bg=bg)
        ws.row_dimensions[ri].height = 16

    # Conditional formatting – highlight interns LWD ≤45 days
    last_row = 2 + len(india_df)
    ws.conditional_formatting.add(
        f"A3:I{last_row}",
        FormulaRule(formula=['$H3="Intern"'], fill=make_fill("FFF3E0"))
    )

    # Status validation
    dv = DataValidation(
        type="list",
        formula1='"Confirmed,Under Probation,Intern"',
        allow_blank=False
    )
    ws.add_data_validation(dv)
    dv.add(f"H3:H{last_row}")

    freeze_and_filter(ws, "A3", f"A2:I{last_row}")
    return ws


def build_us_sheet(wb, us_df):
    if "US Employee Database" in wb.sheetnames:
        del wb["US Employee Database"]
    ws = wb.create_sheet("US Employee Database")
    set_tab_color(ws, C["teal"])

    headers = ["Employee ID","Name","Department","Designation","Reporting Manager",
               "Allocation (%)","CTC (USD)","Skillset","Date of Joining",
               "Employment Status","LWD (Interns)"]
    col_widths = [14, 22, 16, 22, 22, 14, 14, 35, 18, 18, 18]

    ws.merge_cells("A1:K1")
    tc = ws["A1"]
    tc.value = "🇺🇸  US EMPLOYEE DATABASE"
    tc.font = make_font(size=14, bold=True, color=C["white"])
    tc.fill = make_fill(C["navy"])
    tc.alignment = center()
    ws.row_dimensions[1].height = 30

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        style_header_cell(ws.cell(row=2, column=ci), h, bg=C["mid_blue"])
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    today = pd.Timestamp.today()
    for ri, (_, row) in enumerate(us_df.iterrows(), 3):
        bg = C["row_even"] if ri % 2 == 0 else C["row_odd"]
        status = str(row.get("EmploymentStatus",""))
        if status == "Intern":
            lwd = row.get("LWD")
            if pd.notna(lwd) and 0 <= (lwd - today).days <= 45:
                bg = C["alert_red"]
        elif status == "Under Probation":
            bg = C["probation"]

        ctc = row.get("CTC")
        vals = [
            row["EmpID"], row["Name"], row["Department"], row["Designation"],
            row["ReportingManager"],
            f"{int(row['Allocation'])}%" if pd.notna(row.get("Allocation")) else "",
            f"${ctc:,.0f}" if pd.notna(ctc) else "",
            row["Skillset"],
            row["DOJ"].strftime("%d-%b-%Y") if pd.notna(row["DOJ"]) else "",
            row["EmploymentStatus"],
            row["LWD"].strftime("%d-%b-%Y") if pd.notna(row.get("LWD")) else "N/A"
        ]
        for ci, val in enumerate(vals, 1):
            style_data_cell(ws.cell(row=ri, column=ci), val, bg=bg)
        ws.row_dimensions[ri].height = 16

    last_row = 2 + len(us_df)
    freeze_and_filter(ws, "A3", f"A2:K{last_row}")
    return ws


def build_rm_sheet(wb):
    raw = pd.read_excel(SRC, sheet_name="RM Data", header=None)
    if "RM Data" in wb.sheetnames:
        del wb["RM Data"]
    ws = wb.create_sheet("RM Data")
    set_tab_color(ws, C["purple"])

    ws.merge_cells("A1:Z1")
    tc = ws["A1"]
    tc.value = "📊  RESOURCE MANAGEMENT — MONTHLY ALLOCATION DATA"
    tc.font = make_font(size=14, bold=True, color=C["white"])
    tc.fill = make_fill(C["navy"])
    tc.alignment = center()
    ws.row_dimensions[1].height = 30

    months = ["Jan-2025","Feb-2025","Mar-2025","Apr-2025","May-2025","Jun-2025",
              "Jul-2025","Aug-2025","Sep-2025","Oct-2025","Nov-2025","Dec-2025"]

    # Write month group headers (row 2) and sub-headers (row 3)
    base_cols = ["Employee ID","Employee Name","Department","Region"]
    for ci, h in enumerate(base_cols, 1):
        style_header_cell(ws.cell(row=2, column=ci), h, bg=C["navy"])
        style_header_cell(ws.cell(row=3, column=ci), "", bg=C["navy"])
        ws.column_dimensions[get_column_letter(ci)].width = 16 if ci > 1 else 12

    month_colors = [C["mid_blue"], C["dark_blue"]]
    for mi, month in enumerate(months):
        start_col = 5 + mi * 2
        end_col = start_col + 1
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        mc = ws.cell(row=2, column=start_col)
        style_header_cell(mc, month, bg=month_colors[mi % 2])
        style_header_cell(ws.cell(row=3, column=start_col), "Project", bg=month_colors[mi % 2])
        style_header_cell(ws.cell(row=3, column=end_col), "Alloc %", bg=month_colors[mi % 2])
        ws.column_dimensions[get_column_letter(start_col)].width = 18
        ws.column_dimensions[get_column_letter(end_col)].width = 9

    # Write actual data from rows index 2 onwards (skip merged header rows)
    data_rows = raw.iloc[2:].reset_index(drop=True)
    for ri, row in data_rows.iterrows():
        excel_row = ri + 4
        bg = C["row_even"] if ri % 2 == 0 else C["row_odd"]
        for ci in range(1, min(29, len(row) + 1)):
            val = row.iloc[ci - 1]
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=excel_row, column=ci)
            style_data_cell(cell, val, bg=bg)
        ws.row_dimensions[excel_row].height = 15

    ws.freeze_panes = "E4"
    return ws


def build_finance_sheet(wb, finance_df):
    if "Finance" in wb.sheetnames:
        del wb["Finance"]
    ws = wb.create_sheet("Finance")
    set_tab_color(ws, C["green"])

    ws.merge_cells("A1:H1")
    tc = ws["A1"]
    tc.value = "💰  FINANCE — CTC DATA (INDIA & US)"
    tc.font = make_font(size=14, bold=True, color=C["white"])
    tc.fill = make_fill(C["navy"])
    tc.alignment = center()
    ws.row_dimensions[1].height = 30

    headers = ["Employee ID","Name","Department","Region",
               "Annual CTC (INR)","Monthly CTC (INR)","Annual CTC (USD)","Monthly CTC (USD)"]
    widths = [14, 22, 16, 10, 20, 20, 20, 20]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        style_header_cell(ws.cell(row=2, column=ci), h, bg=C["kpi_bg2"])
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    for ri, (_, row) in enumerate(finance_df.iterrows(), 3):
        bg = C["row_even"] if ri % 2 == 0 else C["row_odd"]
        vals = [
            row["EmpID"], row["Name"], row["Department"], row["Region"],
            row["AnnualINR"], row["MonthlyINR"],
            row["AnnualUSD"], row["MonthlyUSD"]
        ]
        fmts = [None, None, None, None, "₹#,##0", "₹#,##0", "$#,##0", "$#,##0"]
        for ci, (val, fmt) in enumerate(zip(vals, fmts), 1):
            style_data_cell(ws.cell(row=ri, column=ci), val, bg=bg,
                            number_format=fmt, align="right" if ci > 4 else "left")
        ws.row_dimensions[ri].height = 16

    # Summary row
    last_row = 2 + len(finance_df)
    sr = last_row + 2
    ws.merge_cells(f"A{sr}:D{sr}")
    style_header_cell(ws.cell(row=sr, column=1), "TOTALS", bg=C["navy"])
    ws.cell(row=sr, column=5).value = f"=SUM(E3:E{last_row})"
    ws.cell(row=sr, column=6).value = f"=SUM(F3:F{last_row})"
    ws.cell(row=sr, column=7).value = f"=SUM(G3:G{last_row})"
    ws.cell(row=sr, column=8).value = f"=SUM(H3:H{last_row})"
    for ci in range(5, 9):
        c = ws.cell(row=sr, column=ci)
        c.font = make_font(bold=True, color=C["white"])
        c.fill = make_fill(C["kpi_bg2"])
        c.border = make_border()
        c.alignment = right_align()

    freeze_and_filter(ws, "A3", f"A2:H{last_row}")
    return ws


def build_productivity_sheet(wb, prod_df):
    if "Productivity" in wb.sheetnames:
        del wb["Productivity"]
    ws = wb.create_sheet("Productivity")
    set_tab_color(ws, C["amber"])

    ws.merge_cells("A1:P1")
    tc = ws["A1"]
    tc.value = "⚡  PRODUCTIVITY — AVERAGE HOURS/DAY PER RESOURCE"
    tc.font = make_font(size=14, bold=True, color=C["white"])
    tc.fill = make_fill(C["navy"])
    tc.alignment = center()
    ws.row_dimensions[1].height = 30

    months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    headers = ["Employee ID","Name","Department","Region"] + \
              [f"{m}-2025" for m in months] + ["Overall Avg","⚠ Flag"]
    widths = [12, 22, 16, 10] + [9]*12 + [12, 16]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        style_header_cell(ws.cell(row=2, column=ci), h, bg=C["dark_gray"])
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    for ri, (_, row) in enumerate(prod_df.iterrows(), 3):
        flag = str(row.get("BelowFlag",""))
        bg = C["alert_red"] if "Below" in flag else (C["row_even"] if ri % 2 == 0 else C["row_odd"])
        vals = [row["EmpID"], row["Name"], row["Department"], row["Region"]]
        for i in range(1, 13):
            vals.append(row.get(f"M{i}", ""))
        vals += [row.get("OverallAvg",""), flag]

        for ci, val in enumerate(vals, 1):
            fmt = "0.0" if ci >= 5 else None
            style_data_cell(ws.cell(row=ri, column=ci), val, bg=bg,
                            number_format=fmt, align="right" if ci >= 5 else "left")
        ws.row_dimensions[ri].height = 15

    # Color scale on monthly data
    last_row = 2 + len(prod_df)
    ws.conditional_formatting.add(
        f"E3:P{last_row}",
        ColorScaleRule(
            start_type="num", start_value=6, start_color="EF5350",
            mid_type="num", mid_value=8, mid_color="FFF9C4",
            end_type="num", end_value=10, end_color="00C853"
        )
    )

    freeze_and_filter(ws, "E3", f"A2:Q{last_row}")
    return ws


def build_risk_sheet(wb, risk_df):
    if "Risk Report" in wb.sheetnames:
        del wb["Risk Report"]
    ws = wb.create_sheet("Risk Report")
    set_tab_color(ws, C["red"])

    ws.merge_cells("A1:J1")
    tc = ws["A1"]
    tc.value = "🚨  RISK REPORT — MAINTAINED BY HR"
    tc.font = make_font(size=14, bold=True, color=C["white"])
    tc.fill = make_fill(C["kpi_bg4"])
    tc.alignment = center()
    ws.row_dimensions[1].height = 30

    headers = ["Employee ID","Name","Department","Region","Risk Category",
               "Risk Level","Identified Date","Mitigation Action","Status","HR Notes"]
    widths = [12, 22, 14, 10, 22, 12, 16, 26, 14, 30]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        style_header_cell(ws.cell(row=2, column=ci), h, bg=C["kpi_bg4"])
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    risk_colors = {"High": "FFCDD2", "Medium": "FFF9C4", "Low": "E8F5E9"}
    for ri, (_, row) in enumerate(risk_df.iterrows(), 3):
        level = str(row.get("RiskLevel",""))
        bg = risk_colors.get(level, C["row_odd"])
        vals = [
            row["EmpID"], row["Name"], row["Department"], row["Region"],
            row["RiskCategory"], row["RiskLevel"],
            row["IdentifiedDate"].strftime("%d-%b-%Y") if pd.notna(row.get("IdentifiedDate")) else "",
            row["MitigationAction"], row["Status"], row["HRNotes"]
        ]
        for ci, val in enumerate(vals, 1):
            bold = ci == 6  # Bold risk level
            style_data_cell(ws.cell(row=ri, column=ci), val, bg=bg, bold=bold)
        ws.row_dimensions[ri].height = 18

    # Status validation
    last_row = 2 + len(risk_df)
    dv_status = DataValidation(type="list", formula1='"Open,In Progress,Resolved,Escalated"')
    dv_level = DataValidation(type="list", formula1='"High,Medium,Low"')
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_level)
    dv_status.add(f"I3:I{last_row}")
    dv_level.add(f"F3:F{last_row}")

    freeze_and_filter(ws, "A3", f"A2:J{last_row}")
    return ws


def build_offboarded_sheet(wb, off_df):
    if "Offboarded Resources" in wb.sheetnames:
        del wb["Offboarded Resources"]
    ws = wb.create_sheet("Offboarded Resources")
    set_tab_color(ws, C["dark_gray"])

    ws.merge_cells("A1:K1")
    tc = ws["A1"]
    tc.value = "🚪  OFFBOARDED RESOURCES — EXIT TRACKING"
    tc.font = make_font(size=14, bold=True, color=C["white"])
    tc.fill = make_fill(C["dark_gray"])
    tc.alignment = center()
    ws.row_dimensions[1].height = 30

    headers = ["Employee ID","Name","Department","Region","Designation",
               "Date of Joining","Last Working Day","Exit Reason",
               "Exit Quarter","Notice Period","Rehire Eligible"]
    widths = [12, 22, 14, 10, 22, 16, 16, 22, 12, 15, 14]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        style_header_cell(ws.cell(row=2, column=ci), h, bg=C["dark_gray"])
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 22

    exit_colors = {
        "Resignation": "FFCDD2",
        "Performance": "FFCDD2",
        "Better Opportunity": "FFF9C4",
        "Personal Reasons": "F3E5F5",
        "Internship Completed": "E8F5E9",
        "End of Contract": "E3F2FD",
        "Relocation": "FFF9C4",
        "Higher Studies": "E8F5E9",
    }

    for ri, (_, row) in enumerate(off_df.iterrows(), 3):
        reason = str(row.get("ExitReason",""))
        bg = exit_colors.get(reason, C["row_odd"])
        vals = [
            row["EmpID"], row["Name"], row["Department"], row["Region"],
            row["Designation"],
            row["DOJ"].strftime("%d-%b-%Y") if pd.notna(row.get("DOJ")) else "",
            row["LWD"].strftime("%d-%b-%Y") if pd.notna(row.get("LWD")) else "",
            row["ExitReason"], row["ExitQuarter"], row["NoticePeriodServed"], row["RehireEligible"]
        ]
        for ci, val in enumerate(vals, 1):
            style_data_cell(ws.cell(row=ri, column=ci), val, bg=bg)
        ws.row_dimensions[ri].height = 16

    last_row = 2 + len(off_df)
    freeze_and_filter(ws, "A3", f"A2:K{last_row}")
    return ws


# ─────────────────────────────────────────────────────────
#  DASHBOARD SHEET (Main)
# ─────────────────────────────────────────────────────────

def write_kpi_card(ws, start_row, start_col, title, value, subtitle, bg_color, icon=""):
    # Merge 3 cols x 5 rows per card
    end_col = start_col + 2
    ws.merge_cells(start_row=start_row, start_column=start_col,
                   end_row=start_row, end_column=end_col)
    ws.merge_cells(start_row=start_row+1, start_column=start_col,
                   end_row=start_row+2, end_column=end_col)
    ws.merge_cells(start_row=start_row+3, start_column=start_col,
                   end_row=start_row+3, end_column=end_col)
    ws.merge_cells(start_row=start_row+4, start_column=start_col,
                   end_row=start_row+4, end_column=end_col)

    title_c = ws.cell(row=start_row, column=start_col)
    title_c.value = f"{icon} {title}"
    title_c.font = make_font(size=9, bold=False, color="B0BEC5")
    title_c.fill = make_fill(bg_color)
    title_c.alignment = center()

    val_c = ws.cell(row=start_row+1, column=start_col)
    val_c.value = value
    val_c.font = make_font(size=22, bold=True, color=C["white"])
    val_c.fill = make_fill(bg_color)
    val_c.alignment = center()

    sub_c = ws.cell(row=start_row+3, column=start_col)
    sub_c.value = subtitle
    sub_c.font = make_font(size=8, color="CFD8DC")
    sub_c.fill = make_fill(bg_color)
    sub_c.alignment = center()

    # Bottom padding row
    pad = ws.cell(row=start_row+4, column=start_col)
    pad.fill = make_fill(bg_color)

    # Set row heights
    for r in range(start_row, start_row+5):
        ws.row_dimensions[r].height = 18


def build_dashboard_sheet(wb, kpis):
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    ws = wb.create_sheet("Dashboard", 0)
    set_tab_color(ws, C["accent"])

    # Set column widths (24 columns for layout)
    for ci in range(1, 25):
        ws.column_dimensions[get_column_letter(ci)].width = 9

    # ── HEADER BANNER ────────────────────────────────────
    ws.merge_cells("A1:X2")
    banner = ws["A1"]
    banner.value = "🏢  HR AUTOMATION DASHBOARD  |  INDIA & US GEOGRAPHIES  |  REAL-TIME ANALYTICS"
    banner.font = make_font(size=16, bold=True, color=C["white"])
    banner.fill = make_fill(C["navy"])
    banner.alignment = center()
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22

    # Sub-header with timestamp note
    ws.merge_cells("A3:X3")
    sub = ws["A3"]
    sub.value = f"Auto-generated from live data tabs  |  Dashboard reflects current state as of: {datetime.today().strftime('%d %B %Y')}  |  ⚠ Red alerts require immediate HR action"
    sub.font = make_font(size=9, italic=True, color=C["mid_gray"])
    sub.fill = make_fill(C["dark_blue"])
    sub.alignment = center()
    ws.row_dimensions[3].height = 16

    # ── SECTION: KPI CARDS (row 5–9) ────────────────────
    ws.merge_cells("A4:X4")
    sec = ws["A4"]
    sec.value = "📊  KEY METRICS AT A GLANCE"
    sec.font = make_font(size=10, bold=True, color=C["white"])
    sec.fill = make_fill(C["mid_blue"])
    sec.alignment = left_align()
    ws.row_dimensions[4].height = 20

    kpi_cards = [
        ("Total Headcount",  kpis["total_hc"],       "India + US Active",  C["kpi_bg1"], "👥", 1),
        ("India Employees",  kpis["india_hc"],        "Active in India",    C["kpi_bg2"], "🇮🇳", 4),
        ("US Employees",     kpis["us_hc"],           "Active in US",       C["kpi_bg3"], "🇺🇸", 7),
        ("On Probation",     kpis["on_probation"],    "Pending Confirmation",C["kpi_bg4"],"⏳", 10),
        ("Active Interns",   kpis["interns"],         "Internship Period",  "1A237E",     "🎓", 13),
        ("Intern LWD Alerts",len(kpis["intern_alerts"]),"Due ≤ 45 Days",   "BF360C",     "🚨", 16),
        ("Risk Cases Open",  len(kpis["risk_dist"]),  "Total Risk Records", "880E4F",     "⚠️", 19),
        ("Avg Productivity", f'{kpis["avg_prod"]:.1f}h',"Avg Hrs/Day All Staff","004D40","⚡", 22),
    ]
    for title, val, sub_lbl, bg, icon, start_col in kpi_cards:
        write_kpi_card(ws, 5, start_col, title, str(val), sub_lbl, bg, icon)

    # ── SECTION LABELS ───────────────────────────────────
    row_sep = 11
    for section, cols, bg in [
        ("🔔  ACTIVE ALERTS", "A11:L11", C["kpi_bg4"]),
        ("📈  DEPARTMENT ANALYTICS", "M11:X11", C["kpi_bg1"]),
    ]:
        ws.merge_cells(cols)
        c = ws[cols.split(":")[0]]
        c.value = section
        c.font = make_font(size=10, bold=True, color=C["white"])
        c.fill = make_fill(bg)
        c.alignment = left_align()
        ws.row_dimensions[row_sep].height = 20

    # ── INTERN LWD ALERT TABLE (rows 12–28) ──────────────
    ws.merge_cells("A12:L12")
    intern_title = ws["A12"]
    intern_title.value = "🚨  INTERN LWD ALERTS — LAST WORKING DAY WITHIN 45 DAYS"
    intern_title.font = make_font(size=9, bold=True, color=C["white"])
    intern_title.fill = make_fill(C["kpi_bg4"])
    intern_title.alignment = center()
    ws.row_dimensions[12].height = 18

    # Sub-headers
    alert_sub = ["Emp ID","Name","Dept","Region","LWD Date","Days Left"]
    alert_cols = [1, 2, 4, 6, 8, 10]
    sub_spans = [1, 2, 2, 2, 2, 2]
    for h, col, span in zip(alert_sub, alert_cols, sub_spans):
        ws.merge_cells(start_row=13, start_column=col, end_row=13, end_column=col+span-1)
        c = ws.cell(row=13, column=col)
        c.value = h
        c.font = make_font(size=8, bold=True, color=C["white"])
        c.fill = make_fill("7B1FA2")
        c.alignment = center()
        c.border = make_border()
    ws.row_dimensions[13].height = 16

    ia_df = kpis["intern_alerts"]
    for ri, (_, row) in enumerate(ia_df.iterrows(), 14):
        days_left = int(row.get("DaysLeft", 0))
        bg = C["alert_red"] if days_left <= 14 else C["alert_amber"]
        data_rows_alerts = [
            (1, 1, row["EmpID"]),
            (2, 2, row["Name"]),
            (4, 2, row["Department"]),
            (6, 2, row.get("Region","India")),
            (8, 2, row["LWD"].strftime("%d-%b-%Y") if pd.notna(row.get("LWD")) else ""),
            (10, 2, f"⚠ {days_left} days"),
        ]
        for col, span, val in data_rows_alerts:
            ws.merge_cells(start_row=ri, start_column=col, end_row=ri, end_column=col+span-1)
            c = ws.cell(row=ri, column=col)
            c.value = val
            c.font = make_font(size=8, bold=(col==10), color="B71C1C" if col==10 else "1A1A1A")
            c.fill = make_fill(bg)
            c.alignment = center()
            c.border = make_border()
        ws.row_dimensions[ri].height = 15
        if ri > 26:
            break

    # No alerts row if empty
    if ia_df.empty:
        ws.merge_cells("A14:L14")
        nc = ws["A14"]
        nc.value = "✅  No intern LWD alerts at this time"
        nc.font = make_font(size=9, color="2E7D32")
        nc.fill = make_fill("E8F5E9")
        nc.alignment = center()
        ws.row_dimensions[14].height = 18

    # ── PROBATION ALERT TABLE (rows 28–40) ───────────────
    ws.merge_cells("A28:L28")
    prob_title = ws["A28"]
    prob_title.value = "⏳  PROBATION CONFIRMATION ALERTS — DUE WITHIN 30 DAYS"
    prob_title.font = make_font(size=9, bold=True, color=C["white"])
    prob_title.fill = make_fill(C["kpi_bg1"])
    prob_title.alignment = center()
    ws.row_dimensions[28].height = 18

    # Sub-headers probation
    for h, col, span in zip(alert_sub, alert_cols, sub_spans):
        ws.merge_cells(start_row=29, start_column=col, end_row=29, end_column=col+span-1)
        c = ws.cell(row=29, column=col)
        c.value = h.replace("LWD Date","Probation Due")
        c.font = make_font(size=8, bold=True, color=C["white"])
        c.fill = make_fill(C["kpi_bg1"])
        c.alignment = center()
        c.border = make_border()
    ws.row_dimensions[29].height = 16

    pa_df = kpis["prob_alerts"]
    for ri, (_, row) in enumerate(pa_df.iterrows(), 30):
        days_left = int(row.get("DaysLeft", 0))
        bg = C["probation"]
        prob_due = row.get("ProbationDue")
        data_rows_prob = [
            (1, 1, row["EmpID"]),
            (2, 2, row["Name"]),
            (4, 2, row["Department"]),
            (6, 2, row.get("Region","India")),
            (8, 2, prob_due.strftime("%d-%b-%Y") if pd.notna(prob_due) else ""),
            (10, 2, f"⏳ {days_left} days"),
        ]
        for col, span, val in data_rows_prob:
            ws.merge_cells(start_row=ri, start_column=col, end_row=ri, end_column=col+span-1)
            c = ws.cell(row=ri, column=col)
            c.value = val
            c.font = make_font(size=8, bold=(col==10), color="0D47A1" if col==10 else "1A1A1A")
            c.fill = make_fill(bg)
            c.alignment = center()
            c.border = make_border()
        ws.row_dimensions[ri].height = 15
        if ri > 39:
            break

    if pa_df.empty:
        ws.merge_cells("A30:L30")
        nc = ws["A30"]
        nc.value = "✅  No probation confirmation alerts at this time"
        nc.font = make_font(size=9, color="2E7D32")
        nc.fill = make_fill("E8F5E9")
        nc.alignment = center()
        ws.row_dimensions[30].height = 18

    # ── DEPT HEADCOUNT TABLE (M12:X28) ───────────────────
    ws.merge_cells("M12:X12")
    dept_title = ws["M12"]
    dept_title.value = "🏬  HEADCOUNT BY DEPARTMENT"
    dept_title.font = make_font(size=9, bold=True, color=C["white"])
    dept_title.fill = make_fill(C["kpi_bg1"])
    dept_title.alignment = center()
    ws.row_dimensions[12].height = 18

    dept_headers = ["Department","Headcount","Share %"]
    dept_col_starts = [13, 19, 22]
    dept_col_spans = [6, 3, 3]
    for h, col, span in zip(dept_headers, dept_col_starts, dept_col_spans):
        ws.merge_cells(start_row=13, start_column=col, end_row=13, end_column=col+span-1)
        c = ws.cell(row=13, column=col)
        c.value = h
        c.font = make_font(size=8, bold=True, color=C["white"])
        c.fill = make_fill(C["dark_blue"])
        c.alignment = center()
        c.border = make_border()

    dept_df = kpis["dept_hc"]
    total = dept_df["Headcount"].sum()
    dept_bar_colors = [C["accent"], C["teal"], C["green"], C["amber"], C["orange"],
                       C["purple"], "EC407A", "26A69A", "FF7043", "AB47BC"]
    for ri, (_, row) in enumerate(dept_df.iterrows(), 14):
        bg = dept_bar_colors[ri % len(dept_bar_colors)]
        pct = f"{row['Headcount']/total*100:.1f}%"
        for col, span, val in [
            (13, 6, row["Department"]),
            (19, 3, row["Headcount"]),
            (22, 3, pct),
        ]:
            ws.merge_cells(start_row=ri, start_column=col, end_row=ri, end_column=col+span-1)
            c = ws.cell(row=ri, column=col)
            c.value = val
            c.font = make_font(size=8, bold=True, color=C["white"])
            c.fill = make_fill("1E3A5F" if col==13 else bg)
            c.alignment = center() if col > 13 else left_align()
            c.border = make_border()
        ws.row_dimensions[ri].height = 15
        if ri > 26:
            break

    # ── ATTRITION SECTION (row 41–55) ────────────────────
    ws.merge_cells("A41:L41")
    att_title = ws["A41"]
    att_title.value = "📉  QUARTERLY ATTRITION RATE  (Auto-Calculated from Offboarded Resources)"
    att_title.font = make_font(size=9, bold=True, color=C["white"])
    att_title.fill = make_fill("33691E")
    att_title.alignment = center()
    ws.row_dimensions[41].height = 18

    att_hdrs = ["Quarter","Total Exits","Attrition Rate %"]
    att_col_starts = [1, 5, 9]
    att_col_spans = [4, 4, 4]
    for h, col, span in zip(att_hdrs, att_col_starts, att_col_spans):
        ws.merge_cells(start_row=42, start_column=col, end_row=42, end_column=col+span-1)
        c = ws.cell(row=42, column=col)
        c.value = h
        c.font = make_font(size=8, bold=True, color=C["white"])
        c.fill = make_fill("558B2F")
        c.alignment = center()
        c.border = make_border()
    ws.row_dimensions[42].height = 16

    att_df = kpis["attrition_q"]
    for ri, (_, row) in enumerate(att_df.iterrows(), 43):
        rate = row["AttritionRate"]
        bg = C["alert_red"] if rate > 10 else (C["alert_amber"] if rate > 5 else "E8F5E9")
        for col, span, val in [
            (1, 4, row["ExitQuarter"]),
            (5, 4, row["Exits"]),
            (9, 4, f"{rate:.1f}%"),
        ]:
            ws.merge_cells(start_row=ri, start_column=col, end_row=ri, end_column=col+span-1)
            c = ws.cell(row=ri, column=col)
            c.value = val
            c.font = make_font(size=9, bold=(col==9), color="B71C1C" if (col==9 and rate>10) else "1A1A1A")
            c.fill = make_fill(bg)
            c.alignment = center()
            c.border = make_border()
        ws.row_dimensions[ri].height = 16

    # ── RISK DISTRIBUTION (M41:X55) ──────────────────────
    ws.merge_cells("M41:X41")
    risk_title = ws["M41"]
    risk_title.value = "🚨  RISK DISTRIBUTION (Live from Risk Report Tab)"
    risk_title.font = make_font(size=9, bold=True, color=C["white"])
    risk_title.fill = make_fill(C["kpi_bg4"])
    risk_title.alignment = center()
    ws.row_dimensions[41].height = 18

    risk_level_colors = {"High": "FFCDD2", "Medium": "FFF9C4", "Low": "E8F5E9"}
    risk_level_fg = {"High": "B71C1C", "Medium": "F57F17", "Low": "2E7D32"}
    risk_df = kpis["risk_dist"]
    for col, span, h in [(13,6,"Risk Level"),(19,3,"Cases"),(22,3,"Status")]:
        ws.merge_cells(start_row=42, start_column=col, end_row=42, end_column=col+span-1)
        c = ws.cell(row=42, column=col)
        c.value = h
        c.font = make_font(size=8, bold=True, color=C["white"])
        c.fill = make_fill(C["kpi_bg4"])
        c.alignment = center()
        c.border = make_border()

    risk_full = load_risk()
    for level in ["High", "Medium", "Low"]:
        count = risk_full[risk_full["RiskLevel"] == level].shape[0]
        open_count = risk_full[(risk_full["RiskLevel"] == level) & (risk_full["Status"] != "Resolved")].shape[0]
        ri = {"High": 43, "Medium": 44, "Low": 45}[level]
        bg = risk_level_colors.get(level, C["row_odd"])
        fg = risk_level_fg.get(level, "1A1A1A")
        for col, span, val in [(13,6,level),(19,3,count),(22,3,f"{open_count} Open")]:
            ws.merge_cells(start_row=ri, start_column=col, end_row=ri, end_column=col+span-1)
            c = ws.cell(row=ri, column=col)
            c.value = val
            c.font = make_font(size=9, bold=True, color=fg)
            c.fill = make_fill(bg)
            c.alignment = center()
            c.border = make_border()
        ws.row_dimensions[ri].height = 18

    # ── FINANCE SUMMARY (row 57–65) ───────────────────────
    ws.merge_cells("A57:X57")
    fin_title = ws["A57"]
    fin_title.value = "💰  FINANCE SUMMARY  |  Total CTC Budget"
    fin_title.font = make_font(size=9, bold=True, color=C["white"])
    fin_title.fill = make_fill(C["kpi_bg2"])
    fin_title.alignment = center()
    ws.row_dimensions[57].height = 18

    fin_cards = [
        ("Total Annual CTC (INR)", f"₹{kpis['total_ctc_inr']:,.0f}", "A58:F62", C["kpi_bg2"]),
        ("Total Annual CTC (USD)", f"${kpis['total_ctc_usd']:,.0f}", "G58:L62", "00695C"),
        ("India Headcount", str(kpis["india_hc"]), "M58:R62", C["kpi_bg1"]),
        ("US Headcount", str(kpis["us_hc"]), "S58:X62", C["kpi_bg3"]),
    ]
    for label, val, merge_range, bg in fin_cards:
        ws.merge_cells(merge_range)
        sc = merge_range.split(":")[0]
        c = ws[sc]
        c.value = f"{label}\n{val}"
        c.font = make_font(size=11, bold=True, color=C["white"])
        c.fill = make_fill(bg)
        c.alignment = center()
        for r in range(58, 63):
            ws.row_dimensions[r].height = 16

    # ── FOOTER ───────────────────────────────────────────
    ws.merge_cells("A64:X64")
    footer = ws["A64"]
    footer.value = ("HR AUTOMATION DASHBOARD  |  Built with Python + openpyxl  "
                    "|  Data Sources: India DB, US DB, RM Data, Finance, Productivity, Risk, Offboarded  "
                    "|  🔴 Red = Urgent  🟡 Amber = Warning  🟢 Green = OK")
    footer.font = make_font(size=7, italic=True, color="78909C")
    footer.fill = make_fill(C["navy"])
    footer.alignment = center()
    ws.row_dimensions[64].height = 14

    # Freeze top rows
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False

    return ws


# ─────────────────────────────────────────────────────────
#  CHART SHEET
# ─────────────────────────────────────────────────────────

def build_charts_sheet(wb, kpis):
    if "Charts & Analytics" in wb.sheetnames:
        del wb["Charts & Analytics"]
    ws = wb.create_sheet("Charts & Analytics")
    set_tab_color(ws, C["accent"])

    ws.merge_cells("A1:P1")
    banner = ws["A1"]
    banner.value = "📊  ANALYTICS CHARTS  |  Auto-Generated from Live Data"
    banner.font = make_font(size=13, bold=True, color=C["white"])
    banner.fill = make_fill(C["navy"])
    banner.alignment = center()
    ws.row_dimensions[1].height = 24
    ws.sheet_view.showGridLines = False

    # ── DATA TABLE: Department Headcount ─────────────────
    ws["A3"] = "Department"
    ws["B3"] = "Headcount"
    ws["A3"].font = make_font(bold=True, color=C["white"])
    ws["A3"].fill = make_fill(C["header_bg"])
    ws["B3"].font = make_font(bold=True, color=C["white"])
    ws["B3"].fill = make_fill(C["header_bg"])
    ws["A3"].alignment = center()
    ws["B3"].alignment = center()

    dept_df = kpis["dept_hc"]
    for ri, (_, row) in enumerate(dept_df.iterrows(), 4):
        ws.cell(row=ri, column=1).value = row["Department"]
        ws.cell(row=ri, column=2).value = row["Headcount"]
    last_dept_row = 3 + len(dept_df)

    # Bar Chart: Dept Headcount
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Headcount by Department"
    bar.y_axis.title = "Employees"
    bar.x_axis.title = "Department"
    bar.height = 12
    bar.width = 18
    cats = Reference(ws, min_col=1, min_row=4, max_row=last_dept_row)
    data_ref = Reference(ws, min_col=2, min_row=3, max_row=last_dept_row)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats)
    bar.shape = 4
    ws.add_chart(bar, "D3")

    # ── DATA TABLE: Attrition ─────────────────────────────
    att_row_start = last_dept_row + 3
    ws.cell(row=att_row_start, column=1).value = "Quarter"
    ws.cell(row=att_row_start, column=2).value = "Exits"
    ws.cell(row=att_row_start, column=3).value = "Attrition %"
    for ci in range(1, 4):
        c = ws.cell(row=att_row_start, column=ci)
        c.font = make_font(bold=True, color=C["white"])
        c.fill = make_fill("33691E")
        c.alignment = center()

    att_df = kpis["attrition_q"]
    for ri, (_, row) in enumerate(att_df.iterrows(), att_row_start+1):
        ws.cell(row=ri, column=1).value = row["ExitQuarter"]
        ws.cell(row=ri, column=2).value = row["Exits"]
        ws.cell(row=ri, column=3).value = row["AttritionRate"] / 100
        ws.cell(row=ri, column=3).number_format = "0.0%"
    last_att_row = att_row_start + len(att_df)

    # Line Chart: Attrition Rate
    line = LineChart()
    line.title = "Quarterly Attrition Rate (%)"
    line.style = 10
    line.y_axis.title = "Rate %"
    line.x_axis.title = "Quarter"
    line.height = 10
    line.width = 18
    l_cats = Reference(ws, min_col=1, min_row=att_row_start+1, max_row=last_att_row)
    l_data = Reference(ws, min_col=3, min_row=att_row_start, max_row=last_att_row)
    line.add_data(l_data, titles_from_data=True)
    line.set_categories(l_cats)
    ws.add_chart(line, "D22")

    # ── DATA TABLE: Risk Distribution ────────────────────
    risk_row_start = last_att_row + 3
    ws.cell(row=risk_row_start, column=1).value = "Risk Level"
    ws.cell(row=risk_row_start, column=2).value = "Count"
    for ci in range(1, 3):
        c = ws.cell(row=risk_row_start, column=ci)
        c.font = make_font(bold=True, color=C["white"])
        c.fill = make_fill(C["kpi_bg4"])
        c.alignment = center()

    risk_full = load_risk()
    risk_counts = risk_full.groupby("RiskLevel").size().reset_index(name="Count")
    for ri, (_, row) in enumerate(risk_counts.iterrows(), risk_row_start+1):
        ws.cell(row=ri, column=1).value = row["RiskLevel"]
        ws.cell(row=ri, column=2).value = row["Count"]
    last_risk_row = risk_row_start + len(risk_counts)

    # Pie Chart: Risk Distribution
    pie = PieChart()
    pie.title = "Risk Level Distribution"
    pie.style = 10
    pie.height = 10
    pie.width = 12
    p_labels = Reference(ws, min_col=1, min_row=risk_row_start+1, max_row=last_risk_row)
    p_data = Reference(ws, min_col=2, min_row=risk_row_start, max_row=last_risk_row)
    pie.add_data(p_data, titles_from_data=True)
    pie.set_categories(p_labels)
    ws.add_chart(pie, "D35")

    # Column widths
    for ci, w in enumerate([18, 12, 14], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    return ws


# ─────────────────────────────────────────────────────────
#  VBA MODULE SHEET (Instructions + VBA Code)
# ─────────────────────────────────────────────────────────

def build_vba_instructions_sheet(wb):
    if "VBA & Setup Guide" in wb.sheetnames:
        del wb["VBA & Setup Guide"]
    ws = wb.create_sheet("VBA & Setup Guide")
    set_tab_color(ws, "9C27B0")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 100

    ws.merge_cells("A1:B1")
    banner = ws["A1"]
    banner.value = "⚙️  VBA MACRO CODE & SETUP GUIDE FOR HR DASHBOARD"
    banner.font = make_font(size=14, bold=True, color=C["white"])
    banner.fill = make_fill("4A148C")
    banner.alignment = center()
    ws.row_dimensions[1].height = 28

    content = [
        ("HEADING", "📌 HOW TO ADD VBA MACROS IN EXCEL"),
        ("TEXT", "1. Press ALT + F11 to open the Visual Basic Editor (VBE)"),
        ("TEXT", "2. Click Insert → Module to create a new module"),
        ("TEXT", "3. Paste the VBA code below into the module"),
        ("TEXT", "4. Press F5 or click Run to execute | Save file as .xlsm (Macro-Enabled Workbook)"),
        ("BLANK", ""),
        ("HEADING", "🔄 VBA: AUTO-REFRESH DASHBOARD BUTTON"),
        ("CODE", """Sub RefreshDashboard()
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationAutomatic
    
    ' Force recalculate all sheets
    Dim ws As Worksheet
    For Each ws In ThisWorkbook.Worksheets
        ws.Calculate
    Next ws
    
    ' Update dashboard timestamp
    Dim dash As Worksheet
    Set dash = ThisWorkbook.Sheets("Dashboard")
    dash.Range("A3").Value = "Auto-generated from live data tabs  |  Last Refreshed: " & Now()
    
    Application.ScreenUpdating = True
    MsgBox "Dashboard refreshed successfully!", vbInformation, "HR Dashboard"
End Sub"""),
        ("BLANK", ""),
        ("HEADING", "🚨 VBA: HIGHLIGHT INTERN ALERTS"),
        ("CODE", """Sub HighlightInternAlerts()
    Dim ws As Worksheet
    Dim lastRow As Long
    Dim today As Date
    today = Date
    
    For Each ws In Array(ThisWorkbook.Sheets("India Employee Database"), _
                         ThisWorkbook.Sheets("US Employee Database"))
        lastRow = ws.Cells(ws.Rows.Count, "A").End(xlUp).Row
        Dim i As Long
        For i = 3 To lastRow
            Dim status As String
            status = ws.Cells(i, 8).Value  ' Employment Status column
            If status = "Intern" Then
                Dim lwd As Variant
                lwd = ws.Cells(i, 9).Value  ' LWD column
                If IsDate(lwd) Then
                    If CDate(lwd) - today <= 45 And CDate(lwd) >= today Then
                        ws.Rows(i).Interior.Color = RGB(255, 205, 210)  ' Red alert
                    End If
                End If
            End If
        Next i
    Next ws
    MsgBox "Intern LWD alerts highlighted!", vbInformation
End Sub"""),
        ("BLANK", ""),
        ("HEADING", "⏳ VBA: FLAG PROBATION EMPLOYEES"),
        ("CODE", """Sub FlagProbationAlerts()
    Dim ws As Worksheet
    Dim lastRow As Long
    Dim today As Date
    today = Date
    
    For Each ws In Array(ThisWorkbook.Sheets("India Employee Database"), _
                         ThisWorkbook.Sheets("US Employee Database"))
        lastRow = ws.Cells(ws.Rows.Count, "A").End(xlUp).Row
        Dim i As Long
        For i = 3 To lastRow
            Dim status As String
            status = ws.Cells(i, 8).Value
            If status = "Under Probation" Then
                Dim doj As Variant
                doj = ws.Cells(i, 7).Value  ' DOJ column
                If IsDate(doj) Then
                    Dim probDue As Date
                    probDue = DateAdd("d", 90, CDate(doj))
                    If probDue - today <= 30 And probDue >= today Then
                        ws.Rows(i).Interior.Color = RGB(227, 242, 253)  ' Blue alert
                    End If
                End If
            End If
        Next i
    Next ws
    MsgBox "Probation alerts flagged!", vbInformation
End Sub"""),
        ("BLANK", ""),
        ("HEADING", "📊 VBA: ADD NAVIGATION BUTTONS TO DASHBOARD"),
        ("CODE", """Sub AddNavigationButtons()
    Dim dash As Worksheet
    Set dash = ThisWorkbook.Sheets("Dashboard")
    
    Dim sheets() As String
    sheets = Array("India Employee Database", "US Employee Database", "RM Data", _
                   "Finance", "Productivity", "Risk Report", "Offboarded Resources", _
                   "Charts & Analytics")
    
    Dim i As Integer
    For i = 0 To UBound(sheets)
        Dim btn As Shape
        Set btn = dash.Shapes.AddShape(msoShapeRoundedRectangle, _
                  10 + i * 95, 5, 88, 24)
        btn.TextFrame.Characters.Text = sheets(i)
        btn.TextFrame.Characters.Font.Size = 7
        btn.TextFrame.Characters.Font.Bold = True
        btn.Fill.ForeColor.RGB = RGB(13, 27, 42)
        btn.TextFrame.Characters.Font.Color = RGB(255, 255, 255)
        btn.Line.Visible = False
    Next i
End Sub"""),
        ("BLANK", ""),
        ("HEADING", "🔐 VBA: PROTECT FORMULA CELLS"),
        ("CODE", """Sub ProtectFormulas()
    Dim ws As Worksheet
    For Each ws In ThisWorkbook.Worksheets
        ws.Unprotect
        ws.Cells.Locked = False
        Dim cell As Range
        For Each cell In ws.UsedRange
            If cell.HasFormula Then
                cell.Locked = True
            End If
        Next cell
        ws.Protect Password:="HRDash2025", DrawingObjects:=True, _
                   Contents:=True, AllowFiltering:=True, _
                   AllowSorting:=True
    Next ws
    MsgBox "Formula cells protected on all sheets!", vbInformation
End Sub"""),
    ]

    row = 2
    for row_type, text in content:
        ws.row_dimensions[row].height = 20 if row_type != "CODE" else 14
        if row_type == "BLANK":
            row += 1
            continue
        c = ws.cell(row=row, column=2)
        c.value = text
        if row_type == "HEADING":
            c.font = make_font(size=11, bold=True, color=C["white"])
            c.fill = make_fill("6A1B9A")
            c.alignment = left_align()
            ws.row_dimensions[row].height = 22
        elif row_type == "CODE":
            c.font = Font(name="Courier New", size=9, color="E0E0E0")
            c.fill = make_fill("1A1A2E")
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            # Count lines and set height
            lines = text.count("\n") + 1
            ws.row_dimensions[row].height = max(14, lines * 13)
        elif row_type == "TEXT":
            c.font = make_font(size=10, color="1A1A1A")
            c.fill = make_fill("F3E5F5")
            c.alignment = left_align()
        row += 1

    return ws


# ─────────────────────────────────────────────────────────
#  MAIN BUILDER
# ─────────────────────────────────────────────────────────

def build_all():
    print("Loading source data...")
    india = load_india()
    us = load_us()
    offboarded = load_offboarded()
    risk = load_risk()
    productivity = load_productivity()
    finance = load_finance()

    print("Computing KPIs and alerts...")
    kpis = compute_kpis(india, us, offboarded, risk, productivity)

    print("Building workbook...")
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print("  -> Building Dashboard sheet...")
    build_dashboard_sheet(wb, kpis)

    print("  -> Building India Employee Database...")
    build_india_sheet(wb, india)

    print("  -> Building US Employee Database...")
    build_us_sheet(wb, us)

    print("  -> Building RM Data...")
    build_rm_sheet(wb)

    print("  -> Building Finance...")
    build_finance_sheet(wb, finance)

    print("  -> Building Productivity...")
    build_productivity_sheet(wb, productivity)

    print("  -> Building Risk Report...")
    build_risk_sheet(wb, risk)

    print("  -> Building Offboarded Resources...")
    build_offboarded_sheet(wb, offboarded)

    print("  -> Building Charts & Analytics...")
    build_charts_sheet(wb, kpis)

    print("  -> Building VBA & Setup Guide...")
    build_vba_instructions_sheet(wb)

    out_path = f"HR_Dashboard_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out_path)
    print(f"DONE: Dashboard saved to: {out_path}")

    return out_path, kpis


if __name__ == "__main__":
    path, kpis = build_all()
    print(f"\nSUMMARY STATS:")
    print(f"   Total Headcount : {kpis['total_hc']}")
    print(f"   India           : {kpis['india_hc']}")
    print(f"   US              : {kpis['us_hc']}")
    print(f"   On Probation    : {kpis['on_probation']}")
    print(f"   Active Interns  : {kpis['interns']}")
    print(f"   Intern Alerts   : {len(kpis['intern_alerts'])} (LWD <=45 days)")
    print(f"   Probation Alerts: {len(kpis['prob_alerts'])} (Due <=30 days)")
    print(f"   Avg Productivity: {kpis['avg_prod']:.2f} hrs/day")